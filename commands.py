import json
import random
import string

import click
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

import database


@click.command("save-quick")
@click.argument("app", nargs=1)
@click.option("-u", "--username", "--user", prompt=True)
@click.option("-p", "--password", "--pass", prompt=True)
def save_quick(app, username, password):
    repo = database.Repository()
    repo.quick_save((app, username, password))


class ListType(click.ParamType):
    def convert(self, value, param, ctx):
        hosts = str(value).split(" ")
        return tuple(set(hosts))


@click.command("save-default")
@click.option("--app", prompt="Application Name", hidden=True)
@click.option("--hosts", prompt=True, hidden=True, type=ListType(), default=str())
@click.option("--emails", prompt=True, hidden=True, type=ListType(), default=str())
@click.option("--username", prompt="Username", hidden=True)
@click.password_option(hidden=True)
def save_default(app, hosts, emails, username, password):
    repo = database.Repository()

    questions = list()
    loop = click.confirm("Would you like to add a Security Question")

    while loop:
        question = click.prompt("Security Question")
        answer = click.prompt("Answer", hide_input=True)
        questions.append({
            "question": question,
            "answer": answer
        })
        loop = click.confirm("Would you like to add a Security Question")

    info = list()
    loop = click.confirm("Would you like to add other information")

    while loop:
        key = click.prompt("Key")
        value = click.prompt("Value")
        info.append({
            "key": key,
            "value": value
        })
        loop = click.confirm("Would you like to add other information")

    credential = {
        "App": app,
        "Hosts": hosts,
        "Emails": emails,
        "Username": username,
        "Password": password,
        "Questions": questions,
        "Other_info": info
    }
    print(credential)
    repo.save(credential)


@click.command("password")
@click.option("--length", "--len", "-l", default=8, type=int)
@click.option("--no-numbers", "--no-nums", "-nn", is_flag=True, default=False)
@click.option("--no-symbols", "--no-syms", "-ns", is_flag=True, default=False)
def create_password(length, no_numbers, no_symbols):
    """
    Create simple password
    """
    letters = list(string.ascii_letters)
    digits = [str(val) for val in range(10)]
    symbols = [val for val in "`~!@#$%^&*()_-+={[}]|\\:;\"'<,>.?/"]

    total, password = letters, list()
    if not no_numbers:
        total = total + digits
    if not no_symbols:
        total = total + symbols

    for i in range(length):
        password.append(random.choice(total))

    password = "".join(password)
    click.echo(password)


@click.command("adv-password")
@click.option("--length", "--len", "-l", default=8, type=int, prompt="Password Length")
@click.option("--minimum-uppercase", "--min-up", "-mu", type=int, default=1,
              prompt="Minimum number of Uppercase Letters")
@click.option("--minimum-lowercase", "--min-low", "-ml", type=int, default=1,
              prompt="Minimum number of Lowercase Letters")
@click.option("--minimum-numbers", "--min-nums", "-mn", type=int, default=1, prompt="Minimum number of digits")
@click.option("--minimum-symbols", "--min-syms", "-ms", type=int, default=1, prompt="Minimum number of symbols")
def create_advance_password(length, minimum_uppercase, minimum_lowercase, minimum_numbers, minimum_symbols):
    """
    Create advanced password with customization
    """
    uppercase = list(string.ascii_uppercase)
    lowercase = list(string.ascii_lowercase)
    digits = [str(val) for val in range(10)]
    symbols = [val for val in "`~!@#$%^&*()_-+={[}]|\\:;\"'<,>.?/"]

    total = uppercase + lowercase + digits + symbols
    password = list()
    if length < (minimum_uppercase + minimum_lowercase + minimum_symbols):
        click.echo("Invalid input.")
    else:
        for i in range(minimum_uppercase):
            password.append(random.choice(uppercase))
        length = length - minimum_uppercase

        for i in range(minimum_lowercase):
            password.append(random.choice(lowercase))
        length = length - minimum_lowercase

        for i in range(minimum_numbers):
            password.append(random.choice(digits))
        length = length - minimum_numbers

        for i in range(minimum_symbols):
            password.append(random.choice(symbols))
        length = length - minimum_symbols

        for i in range(length):
            password.append(random.choice(total))

        random.shuffle(password)
        password = "".join(password)

        click.echo(password)


@click.command("apps")
def get_apps():
    """
    Get all saved application names
    """
    repo = database.Repository()
    apps = repo.get_all_apps()
    for idx in range(len(apps)):
        click.echo(f"{idx + 1}. {apps[idx][0]}")


@click.command("details")
@click.argument("app")
def get_details(app):
    """
    Get full details of an Application
    """
    repo = database.Repository()
    details = repo.get_details_by_app(app)
    if len(details) != 0:
        click.echo(json.dumps(details, indent=2))


@click.command("password")
@click.argument("app")
def get_password(app):
    """
    Get Password of an Application
    """
    repo = database.Repository()
    password = repo.get_password_by_app(app)
    if len(password) > 0:
        click.echo(password)


@click.command("template")
@click.option("--file", "-f", default="template.xlsx")
def create_template(file):
    """
    Generates a sample Excel template file
    """
    headings = {
        "credentials": ["Application Name", "Username", "Password", "Hosts (comma separated)",
                        "emails (comma separated)"],
        "questions": ["Application Name (Repeatable)", "Security Question", "Answer"],
        "info": ["Application Name", "Key", "Value"]
    }
    workbook = Workbook()
    workbook.remove(workbook.active)
    credentials = workbook.create_sheet(title="Credentials", index=0)
    questions = workbook.create_sheet(title="Security Questions", index=1)
    info = workbook.create_sheet(title="Other Info", index=2)

    for i in range(len(headings.get("credentials"))):
        credentials.cell(row=1, column=i + 1).value = headings.get("credentials")[i]
    for i in range(len(headings.get("questions"))):
        questions.cell(row=1, column=i + 1).value = headings.get("questions")[i]
    for i in range(len(headings.get("info"))):
        info.cell(row=1, column=i + 1).value = headings.get("info")[i]

    workbook.save(filename=file)
    workbook.close()


@click.command("export")
@click.option("--file", "-f", default="exported-data.xlsx")
def export(file):
    """
    Exports all data to a file
    """
    repo = database.Repository()
    workbook = Workbook()
    workbook.remove(workbook.active)
    credentials = workbook.create_sheet(title="Credentials", index=0)
    questions = workbook.create_sheet(title="Security Questions", index=1)
    info = workbook.create_sheet(title="Other Info", index=2)

    data = repo.get_all_data()
    row = 1
    for val in data.values():
        credentials.cell(row=row, column=1).value = val["app"]
        credentials.cell(row=row, column=2).value = val["username"]
        credentials.cell(row=row, column=3).value = val["password"]
        credentials.cell(row=row, column=4).value = " ".join(val["hosts"])
        credentials.cell(row=row, column=5).value = " ".join(val["emails"])
        row += 1

    row = 1
    for item in data.values():
        app, ques_list = item["app"], item["questions"]
        print(ques_list)
        for val in ques_list:
            questions.cell(row=row, column=1).value = app
            questions.cell(row=row, column=2).value = val["question"]
            questions.cell(row=row, column=3).value = val["answer"]
            row += 1

    row = 1
    for item in data.values():
        app, info_list = item["app"], item["info"]
        for val in info_list:
            info.cell(row=row, column=1).value = app
            info.cell(row=row, column=2).value = val["key"]
            info.cell(row=row, column=3).value = val["val"]
            row += 1

    workbook.save(filename=file)
    workbook.close()
